"""Excel output with color coding"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def write_results_to_excel(results, output_path):
    """
    Write screening results to Excel with color coding
    
    Color coding:
    - Red: Percentile <= -10% (extreme negative move)
    - Yellow: -10% to -5% or +5% to +10% (moderate move)
    - Green: Percentile >= +10% (extreme positive move)
    """
    df = pd.DataFrame(results)
    
    # Write to Excel
    df.to_excel(output_path, index=False, sheet_name='Screening Results')
    
    # Apply color coding
    wb = load_workbook(output_path)
    ws = wb['Screening Results']
    
    # Define fills
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    
    # Find percentile columns (p1, p5, p10, p25, p50, p75, p90, p95, p99)
    percentile_cols = []
    for col_idx, col in enumerate(df.columns, start=1):
        if col.startswith('p') and col[1:].isdigit():
            percentile_cols.append(col_idx)
    
    # Apply colors to data rows (skip header)
    for row_idx in range(2, len(df) + 2):
        for col_idx in percentile_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                value = float(cell.value)
                
                if value <= -10:
                    cell.fill = red_fill
                elif value <= -5:
                    cell.fill = yellow_fill
                elif value >= 10:
                    cell.fill = green_fill
                elif value >= 5:
                    cell.fill = yellow_fill
            except (ValueError, TypeError):
                pass
    
    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")